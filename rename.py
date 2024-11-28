import sys
import os
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox

class RenameAppTk(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Batch Rename Files')
        self.geometry('1000x600')
        self.configure(bg='#4b6b63')
        self.initUI()

    def initUI(self):
        # Define a custom style
        self.option_add('*TButton*Font', ('Arial', 12))
        self.option_add('*TButton*Foreground', 'white')
        self.option_add('*TButton*Background', '#eda1f7')
        self.option_add('*TLabel*Font', ('Arial', 12))
        self.option_add('*TLabel*Foreground', 'white')
        self.option_add('*TLabel*Background', '#f7ebeb')
        self.option_add('*TEntry*Font', ('Arial', 12))
        self.option_add('*TEntry*Foreground', 'white')
        self.option_add('*TEntry*Background', '#eda1f7')
        self.option_add('*TText*Font', ('Arial', 12))
        self.option_add('*TText*Foreground', 'white')
        self.option_add('*TText*Background', '#eda1f7')

        # Excel File Path
        self.excel_label = tk.Label(self, text='Excel File Path:', bg='#121212', fg='white', font=('Arial', 12))
        self.excel_label.grid(row=0, column=0, padx=10, pady=10, sticky='e')
        self.excel_path_input = tk.Entry(self, width=70, bg='#333333', fg='white', font=('Arial', 12))
        self.excel_path_input.grid(row=0, column=1, padx=10, pady=10)
        self.excel_browse_button = tk.Button(self, text='Browse', command=self.select_excel_file, bg='#333333', fg='white', font=('Arial', 12))
        self.excel_browse_button.grid(row=0, column=2, padx=10, pady=10)

        # Directory Path
        self.directory_label = tk.Label(self, text='Directory Path:', bg='#121212', fg='white', font=('Arial', 12))
        self.directory_label.grid(row=1, column=0, padx=10, pady=10, sticky='e')
        self.directory_path_input = tk.Entry(self, width=70, bg='#333333', fg='white', font=('Arial', 12))
        self.directory_path_input.grid(row=1, column=1, padx=10, pady=10)
        self.directory_browse_button = tk.Button(self, text='Browse', command=self.select_directory, bg='#333333', fg='white', font=('Arial', 12))
        self.directory_browse_button.grid(row=1, column=2, padx=10, pady=10)

        # Separator Input
        self.separator_label = tk.Label(self, text='File Name Separator:', bg='#121212', fg='white', font=('Arial', 12))
        self.separator_label.grid(row=2, column=0, padx=10, pady=10, sticky='e')
        self.separator_input = tk.Entry(self, width=10, bg='#333333', fg='white', font=('Arial', 12))
        self.separator_input.grid(row=2, column=1, padx=10, pady=10)
        self.separator_input.insert(0, '-')  # Default separator

        # Rename Button
        self.rename_button = tk.Button(self, text='Rename Files', command=self.rename_files, bg='#333333', fg='white', font=('Arial', 12))
        self.rename_button.grid(row=3, column=1, pady=20)

        # Message Display with Help Text
        self.message_text = tk.Text(self, height=15, width=90, bg='#121212', fg='#FFFFFF', font=('Arial', 12))
        self.message_text.grid(row=4, column=0, columnspan=3, padx=10, pady=10)
        self.message_text.config(state=tk.NORMAL)
        self.message_text.insert(tk.END, "请先选择命名格式文件（xls、xlsx），命名格式文件请手动输入信息，之后请选择目标文件夹。目标文件夹中文件数目应和命名文件行数一致，之后输入你希望使用的连接符，默认为“-“，点击rename开始重命名。\n")
        self.message_text.see(tk.END)
        self.message_text.config(state=tk.DISABLED)

    def select_excel_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")])
        if file_path:
            self.excel_path_input.delete(0, tk.END)
            self.excel_path_input.insert(0, file_path)

    def select_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.directory_path_input.delete(0, tk.END)
            self.directory_path_input.insert(0, directory)

    def rename_files(self):
        excel_path = self.excel_path_input.get()
        directory_path = self.directory_path_input.get()
        separator = self.separator_input.get()

        self.message_text.config(state=tk.NORMAL)
        self.message_text.delete('1.0', tk.END)
        self.message_text.insert(tk.END, f'Processing files in directory "{directory_path}" using data from "{excel_path}"\n')

        # Load workbook and get active sheet
        wb = load_workbook(filename=excel_path)
        sheet = wb.active

        # Get all data from the sheet, skipping empty rows
        attributes = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if any(cell for cell in row):  # Skip empty rows
                attributes.append(row)

        # Determine the number of columns dynamically
        num_columns = max(len(row) for row in attributes)
        attributes = [list(row) + [None] * (num_columns - len(row)) for row in attributes]

        # Transpose the data to get columns as lists
        attributes = list(zip(*attributes))

        current_files = os.listdir(directory_path)

        if len(attributes[0]) != len(current_files):
            self.message_text.insert(tk.END, "Error: The number of rows in the Excel does not match the number of files in the directory.\n")
            return

        for index, row in enumerate(zip(*attributes)):
            # Use the user-defined separator when constructing new filenames
            new_filename = separator.join(str(item) for item in row if item is not None) + ".docx"
            old_file = current_files[index]
            old_file_path = os.path.join(directory_path, old_file)
            new_file_path = os.path.join(directory_path, new_filename)

            if not os.path.exists(new_file_path):
                os.rename(old_file_path, new_file_path)
                self.message_text.insert(tk.END, f'Renamed "{old_file}" to "{new_filename}"\n')
            else:
                self.message_text.insert(tk.END, f'Error: The file "{new_filename}" already exists. Skipping rename for "{old_file}".\n')

        self.message_text.insert(tk.END, 'Renaming operation completed successfully.\n')
        self.message_text.see(tk.END)
        self.message_text.config(state=tk.DISABLED)

if __name__ == '__main__':
    app = RenameAppTk()
    app.mainloop()
